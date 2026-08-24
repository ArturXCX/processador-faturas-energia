"""
Processamento de BORDERÔS de energia (faturas AGRUPADAS).

Um borderô é uma *fatura agrupada*: um único documento da distribuidora
(ENEL/CELG ou EQUATORIAL Goiás) que consolida dezenas/centenas de unidades
consumidoras (UCs), cada uma com seu valor, fechando com um TOTAL. Diferente
das faturas individuais (ver `equatorial.py`/`chesp.py`), aqui há poucos PDFs
(≈12/ano) e o objetivo é tabular:

  * o cabeçalho de cada borderô (competência, vencimento, total, nº de contas);
  * o "RESUMO DOS VALORES COBRADOS" (rubricas tarifárias do agrupamento);
  * o detalhamento por UC (a lista que caracteriza o borderô).

Dois layouts foram observados nos dados reais do TJGO:

  ┌────────────┬───────────────────────────────────────────────────────────────┐
  │ Layout     │ Colunas da tabela de UCs                                       │
  ├────────────┼───────────────────────────────────────────────────────────────┤
  │ ENEL/CELG  │ UNIDADE CONSUMIDORA · NOTA FISCAL · VALOR                      │
  │ (até ~2023)│ (1 ou 2 colunas físicas por página)                           │
  ├────────────┼───────────────────────────────────────────────────────────────┤
  │ EQUATORIAL │ (NÚMERO DA) UC · VALOR BRUTO · COFINS · PIS · IRRF · CSLL ·    │
  │ (2023-...) │  VALOR LÍQUIDO                                                 │
  └────────────┴───────────────────────────────────────────────────────────────┘

A extração é POSICIONAL (coordenadas das palavras via PyMuPDF), o que torna o
parser robusto às duas variantes do ENEL e à mudança de rótulo/dígitos da UC na
EQUATORIAL. Cada borderô é conferido por dois âncoras impressos no próprio PDF:
a soma dos valores das UCs deve bater com o TOTAL, e a contagem de UCs com o
"nº de contas" (quando informado). Validação nos 54 PDFs de 2022–2026: 53/54
reconciliam ao centavo; o único fora é fev/2022, cujas páginas de detalhe são
digitalizadas (imagem) e ilegíveis — o cabeçalho é extraído e o borderô é
marcado como `escaneado`.
"""
from __future__ import annotations

import os
import re
import glob
from dataclasses import dataclass, field

import fitz  # PyMuPDF


# --------------------------------------------------------------------------- #
# Esquema das abas geradas
# --------------------------------------------------------------------------- #
BORDERO_COLS = [
    "id_bordero", "numero_fatura_agrupada", "arquivo_pdf", "distribuidora",
    "cod_agrupamento", "competencia", "data_vencimento",
    "quantidade_contas", "quantidade_contas_extraidas",
    "valor_total", "valor_total_bruto", "valor_total_retencoes",
    "soma_valores_extraidos", "bate_total", "escaneado", "observacao",
]
UNIDADE_COLS = [
    "id_bordero", "distribuidora", "competencia", "unidade_consumidora",
    "nota_fiscal", "valor_bruto", "cofins", "pis", "irrf", "csll", "valor",
]
RESUMO_COLS = [
    "id_bordero", "distribuidora", "competencia", "codigo", "descricao", "valor",
]

# Nomes e ordem das abas na planilha final.
ABA_BORDEROS = "borderos"
ABA_UNIDADES = "unidades"
ABA_RESUMO = "resumo_valores"
SHEET_ORDER = [ABA_BORDEROS, ABA_UNIDADES, ABA_RESUMO]

# Paleta (cabeçalho, zebra) por aba — mesmo espírito visual do restante do app.
SHEET_COLORS = {
    ABA_BORDEROS: ("1F4E79", "BDD7EE"),
    ABA_UNIDADES: ("375623", "E2EFDA"),
    ABA_RESUMO:   ("7B2C2C", "FCE4D6"),
}

# Tolerância (R$) para considerar que a soma das UCs bate com o total impresso.
TOL_TOTAL = 1.0


# --------------------------------------------------------------------------- #
# Utilidades de baixo nível
# --------------------------------------------------------------------------- #
_MONEY = re.compile(r"^-?\d{1,3}(?:\.\d{3})*,\d{1,2}$|^-?\d+,\d{1,2}$")
_INT = re.compile(r"^\d{5,15}$")
_CODIGO = re.compile(r"^[A-Z0-9][A-Z0-9\-]{0,11}$")  # PEFP, AB1234, 54, 53-COMP


def _money(s: str) -> float:
    return float(s.replace(".", "").replace(",", "."))


def _fmt_valor(v):
    """Formata float como string monetária pt-BR (1.234,56); vazio se None."""
    if v is None:
        return ""
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _clusterizar_linhas(words, ytol=3):
    """Agrupa as palavras (x0,y0,x1,y1,texto,...) em linhas por proximidade de y."""
    ws = sorted(words, key=lambda t: (round(t[1] / ytol), t[0]))
    linhas, atual, cy = [], [], None
    for t in ws:
        if cy is None or abs(t[1] - cy) <= ytol:
            atual.append(t)
            cy = t[1] if cy is None else cy
        else:
            linhas.append(atual)
            atual, cy = [t], t[1]
    if atual:
        linhas.append(atual)
    return linhas


# --------------------------------------------------------------------------- #
# Detecção de layout
# --------------------------------------------------------------------------- #
def _texto_completo(doc) -> str:
    return "".join(doc[i].get_text() for i in range(doc.page_count))


def _is_layout_equatorial(texto_completo: str) -> bool:
    return "COFINS" in texto_completo or "VALOR BRUTO" in texto_completo


# --------------------------------------------------------------------------- #
# Extração das UCs — layout ENEL/CELG (UC · NF · VALOR)
# --------------------------------------------------------------------------- #
def _extrair_unidades_enel(doc):
    """
    Máquina de estados por linha: acumula inteiros e, ao encontrar um valor
    monetário, fecha um registro (UC = 1º inteiro, NF = 2º inteiro, VALOR = o
    monetário). Reiniciar a cada linha isola o cabeçalho (que não tem valores)
    e cobre tanto o layout de 1 coluna quanto o de 2 colunas físicas.
    """
    regs = []
    for pi in range(1, doc.page_count):
        for linha in _clusterizar_linhas(doc[pi].get_text("words")):
            ints = []
            for t in sorted(linha, key=lambda x: x[0]):
                txt = t[4].strip()
                if _MONEY.match(txt):
                    if len(ints) >= 2:
                        regs.append({
                            "unidade_consumidora": ints[0],
                            "nota_fiscal": ints[1],
                            "valor": _money(txt),
                        })
                    ints = []
                elif _INT.match(txt):
                    ints.append(txt)
                else:
                    ints = []  # token estranho quebra a sequência
    return regs


# --------------------------------------------------------------------------- #
# Extração das UCs — layout EQUATORIAL (7 colunas)
# --------------------------------------------------------------------------- #
# Limite superior de x (pt) de cada coluna -> campo. Colunas são justificadas à
# direita; os limites separam com folga as faixas observadas nos PDFs.
_EQ_BINS = [
    (157, "unidade_consumidora"),
    (242, "valor_bruto"),
    (316, "cofins"),
    (385, "pis"),
    (447, "irrf"),
    (515, "csll"),
    (10_000, "valor_liquido"),
]


def _eq_coluna(x):
    for lim, campo in _EQ_BINS:
        if x < lim:
            return campo
    return "valor_liquido"


def _extrair_unidades_equatorial(doc):
    regs = []
    for pi in range(1, doc.page_count):
        for linha in _clusterizar_linhas(doc[pi].get_text("words")):
            uc = None
            vals = {}
            for t in sorted(linha, key=lambda x: x[0]):
                txt = t[4].strip()
                col = _eq_coluna(t[0])
                if col == "unidade_consumidora":
                    if _INT.match(txt):
                        uc = txt
                elif _MONEY.match(txt):
                    vals[col] = _money(txt)
            if uc and "valor_liquido" in vals:
                regs.append({
                    "unidade_consumidora": uc,
                    "nota_fiscal": "",
                    "valor_bruto": vals.get("valor_bruto"),
                    "cofins": vals.get("cofins"),
                    "pis": vals.get("pis"),
                    "irrf": vals.get("irrf"),
                    "csll": vals.get("csll"),
                    "valor": vals.get("valor_liquido"),
                })
    return regs


# --------------------------------------------------------------------------- #
# RESUMO DOS VALORES COBRADOS (rubricas tarifárias do agrupamento — página 0)
# --------------------------------------------------------------------------- #
def _extrair_resumo(doc):
    """
    Lê a grade CÓDIGO · DESCRIÇÃO · VALOR da página 0. Mesma ideia da ENEL:
    acumula tokens de texto e fecha uma rubrica ao encontrar o valor monetário
    (cobre as 1 ou 2 colunas físicas do resumo).
    """
    regs = []
    for linha in _clusterizar_linhas(doc[0].get_text("words")):
        acc = []
        for t in sorted(linha, key=lambda x: x[0]):
            txt = t[4].strip()
            if _MONEY.match(txt):
                if acc and _CODIGO.match(acc[0]) and acc[0] not in ("CÓDIGO", "C�DIGO"):
                    codigo = acc[0]
                    descricao = " ".join(acc[1:]).strip()
                    if descricao:  # rubrica precisa de descrição
                        regs.append({"codigo": codigo, "descricao": descricao,
                                     "valor": _money(txt)})
                acc = []
            elif re.match(r"^R\$", txt) or txt in ("VALOR", "DESCRIÇÃO", "DESCRI��O"):
                acc = []
            else:
                acc.append(txt)
    return regs


# --------------------------------------------------------------------------- #
# Cabeçalho / âncoras de conferência
# --------------------------------------------------------------------------- #
def _anchor_total(full: str):
    """TOTAL impresso (o valor líquido do agrupamento)."""
    total = None
    for m in re.finditer(r"R\$\*+\s*([\d.]+,\d{2})", full):
        total = _money(m.group(1))  # o último R$*** costuma ser o total final
    m = re.search(r"VALOR L[I�]QUIDO EM REAL\s*R\$\*+\s*([\d.]+,\d{2})", full)
    if m:
        total = _money(m.group(1))
    m = re.search(r"TOTAL:\s*R\$\*+\s*([\d.]+,\d{1,2})", full)
    if m:
        total = _money(m.group(1))
    return total


def _anchor_contas(full: str):
    m = re.search(r"Quantidade de contas[:\s]*([0-9]+)", full)
    if m:
        return int(m.group(1))
    m = re.search(r"([0-9]+)\s*CONTA\(S\)", full)
    if m:
        return int(m.group(1))
    return None


def _anchor_bruto_retencoes(full: str):
    bruto = retencoes = None
    m = re.search(r"VALOR TOTAL BRUTO:\s*([\d.]+,\d{2})", full)
    if m:
        bruto = _money(m.group(1))
    m = re.search(r"VALOR TOTAL RETEN[�Ç][�Õ]ES:\s*(-?[\d.]+,\d{2})", full)
    if m:
        retencoes = _money(m.group(1))
    return bruto, retencoes


def _cod_agrupamento(full: str):
    m = re.search(r"\b(00\d{5})\b", full)
    return m.group(1) if m else ""


# --------------------------------------------------------------------------- #
# Identidade do borderô (número, agrupamento, competência, vencimento)
# --------------------------------------------------------------------------- #
# Antes tudo isso vinha do NOME DO ARQUIVO, com o prefixo `4000000225` — o
# código de agrupamento do TJGO — escrito no regex. Bastava renomear o PDF (ou
# ser de outra instituição) para o borderô ficar sem identidade nenhuma.
#
# Agora a fonte é o CONTEÚDO do PDF, que traz tudo impresso no cabeçalho:
#
#   EQUATORIAL            ENEL/CELG (rótulos e valores em linhas separadas)
#   ----------            ---------------------------------------------------
#   0022500               NÚMERO DA FATURA AGRUPADA
#   08/2024               4000000225001202200
#   30/09/2024            CÓD. AGRUPAMENTO   MÊS REF.   VENCIMENTO   VALOR
#   R$***1.144.838,19     0022500            01/2022    28/02/2022   R$***…
#   4000000225008202400
#
# O nome do arquivo continua valendo como ÚLTIMO recurso (ele é padronizado no
# acervo do TJGO e ajuda quando o PDF é digitalizado e não tem texto).
_RE_COMPETENCIA = re.compile(r"\b(0[1-9]|1[0-2])/(19|20)(\d{2})\b")
_RE_DATA = re.compile(r"\b(\d{2})/(\d{2})/((?:19|20)\d{2})\b")
_RE_NUM_AGRUPADA = re.compile(r"\b(\d{15,25})\b")
_RE_COD_AGRUP = re.compile(r"\b(\d{7})\b")

_MESES_EXT = {"JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "MARÇO": 3, "ABRIL": 4,
              "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9,
              "SET": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12}


def _linhas_pagina(doc, pagina=0):
    """Linhas da página como texto, na ordem visual (usa as coordenadas)."""
    if doc.page_count <= pagina:
        return []
    palavras = doc[pagina].get_text("words")
    return [" ".join(w[4] for w in linha) for linha in _clusterizar_linhas(palavras)]


def _sem_acento_upper(s: str) -> str:
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in s if not unicodedata.combining(c)).upper()


def _distribuidora(full: str, nome: str) -> str:
    """
    Quem emitiu — do conteúdo; o nome do arquivo é o desempate.

    O fallback pelo nome NÃO é preciosismo: em boa parte do acervo (todo 2022 e
    jan–set/2023) a marca da concessionária não aparece no TEXTO do PDF, porque
    o cabeçalho é imagem. Nesses casos o nome do arquivo é a única fonte.

    O layout da tabela seria o candidato natural a substituir isso, mas não
    serve: jun–set/2023 são borderôs da EQUATORIAL ainda emitidos no template
    da ENEL (a concessão trocou antes do documento). Usar o layout daria
    distribuidora errada justamente nesses meses.
    """
    for fonte in (full, nome):
        up = _sem_acento_upper(fonte)
        if "EQUATORIAL" in up:
            return "EQUATORIAL"
        if "ENEL" in up or "CELG" in up:
            return "ENEL"
    return ""


def _competencia_iso(mm, aaaa) -> str:
    return f"{int(aaaa):04d}-{int(mm):02d}"


def _identidade(doc, full: str, nome: str) -> dict:
    """
    Número, código de agrupamento, competência e vencimento — do CONTEÚDO,
    caindo para o nome do arquivo. Competência sai em `AAAA-MM` e vencimento em
    `AAAA-MM-DD`, como no resto do app (antes eram `MM/AAAA` e `dd/mm/aaaa`,
    o que impedia cruzar borderô com fatura numa tabela dinâmica).
    """
    numero = cod = competencia = vencimento = ""
    linhas = _linhas_pagina(doc, 0)

    # 1) Linha de VALORES logo abaixo da linha de RÓTULOS (layout ENEL) ------
    for i, linha in enumerate(linhas):
        up = _sem_acento_upper(linha)
        if "MES REF" in up and "VENCIMENTO" in up:
            for seguinte in linhas[i + 1:i + 3]:
                mc = _RE_COMPETENCIA.search(seguinte)
                md = _RE_DATA.search(seguinte)
                if mc:
                    competencia = _competencia_iso(mc.group(1),
                                                   mc.group(2) + mc.group(3))
                if md:
                    vencimento = f"{md.group(3)}-{md.group(2)}-{md.group(1)}"
                mcod = _RE_COD_AGRUP.search(seguinte)
                if mcod:
                    cod = mcod.group(1)
                if competencia:
                    break
            break
        if "NUMERO DA FATURA AGRUPADA" in up:
            for seguinte in linhas[i + 1:i + 3]:
                mn = _RE_NUM_AGRUPADA.search(seguinte)
                if mn:
                    numero = mn.group(1)
                    break

    # 2) EQUATORIAL: cabeçalho solto no topo, sem rótulos ao lado -----------
    if not competencia or not vencimento or not numero:
        topo = "\n".join(linhas[:12])
        if not competencia:
            mc = _RE_COMPETENCIA.search(topo)
            if mc:
                competencia = _competencia_iso(mc.group(1), mc.group(2) + mc.group(3))
        if not vencimento:
            md = _RE_DATA.search(topo)
            if md:
                vencimento = f"{md.group(3)}-{md.group(2)}-{md.group(1)}"
        if not numero:
            mn = _RE_NUM_AGRUPADA.search(topo)
            if mn:
                numero = mn.group(1)
        if not cod:
            mcod = _RE_COD_AGRUP.search(topo)
            if mcod:
                cod = mcod.group(1)

    # 3) Documento todo (PDF com o cabeçalho em outra página) ---------------
    if not numero:
        mn = _RE_NUM_AGRUPADA.search(full)
        if mn:
            numero = mn.group(1)
    if not cod:
        cod = _cod_agrupamento(full)

    # 4) Último recurso: o nome do arquivo ----------------------------------
    if not numero:
        mn = _RE_NUM_AGRUPADA.search(nome)
        if mn:
            numero = mn.group(1)
    if not competencia:
        # "… - AGOSTO.2024 - …" / "… - SET.2025 - …"
        mm = re.search(r"\b([A-ZÇ]{3,9})\.?\s*((?:19|20)\d{2})\b",
                       _sem_acento_upper(nome).replace("Ç", "C"))
        if mm and mm.group(1) in _MESES_EXT:
            competencia = _competencia_iso(_MESES_EXT[mm.group(1)], mm.group(2))
        else:
            mc = _RE_COMPETENCIA.search(nome)
            if mc:
                competencia = _competencia_iso(mc.group(1), mc.group(2) + mc.group(3))
    if not vencimento:
        mv = re.search(r"VENC\.?\s*(\d{2})\.(\d{2})\.(\d{2,4})", nome, re.IGNORECASE)
        if mv:
            dd, mmv, yy = mv.groups()
            if len(yy) == 2:
                yy = "20" + yy
            vencimento = f"{yy}-{mmv}-{dd}"

    return {"numero": numero, "cod_agrupamento": cod,
            "competencia": competencia, "vencimento": vencimento}


def _montar_id(distribuidora: str, numero: str, competencia: str) -> str:
    """
    `id_bordero`: prefixo da distribuidora + número da fatura agrupada.

    Sem número (PDF digitalizado, layout novo, arquivo renomeado), cai para
    DISTRIBUIDORA_AAAAMM — que é único por borderô, já que cada distribuidora
    emite um por mês. Nunca devolve vazio: sem nada disso, usa o que houver.
    """
    dist = distribuidora or "BORDERO"
    if numero:
        return f"{dist}_{numero}"
    if competencia:
        return f"{dist}_{competencia.replace('-', '')}"
    return dist


# --------------------------------------------------------------------------- #
# API pública
# --------------------------------------------------------------------------- #
@dataclass
class ResultadoBordero:
    bordero: dict
    unidades: list = field(default_factory=list)
    resumo: list = field(default_factory=list)


def processar_pdf(path: str) -> ResultadoBordero:
    """Processa um PDF de borderô e devolve o registro-cabeçalho + detalhes."""
    nome = os.path.basename(path)
    doc = fitz.open(path)
    full = _texto_completo(doc)
    distribuidora = _distribuidora(full, nome)
    ident = _identidade(doc, full, nome)
    numero = ident["numero"]
    competencia = ident["competencia"]
    vencimento = ident["vencimento"]
    id_bordero = _montar_id(distribuidora, numero, competencia)

    total = _anchor_total(full)
    contas = _anchor_contas(full)
    bruto, retencoes = _anchor_bruto_retencoes(full)
    cod_agr = ident["cod_agrupamento"] or _cod_agrupamento(full)

    # Detalhe por UC só é confiável se as páginas têm texto (não são escaneadas).
    texto_detalhe = sum(len(doc[i].get_text().strip()) for i in range(1, doc.page_count))
    escaneado = texto_detalhe < 300

    unidades = []
    resumo = []
    if not escaneado:
        if _is_layout_equatorial(full):
            unidades = _extrair_unidades_equatorial(doc)
        else:
            unidades = _extrair_unidades_enel(doc)
        resumo = _extrair_resumo(doc)

    soma = round(sum(u["valor"] for u in unidades if u.get("valor") is not None), 2)
    bate = (total is not None) and (abs(soma - total) < TOL_TOTAL)

    obs = []
    if escaneado:
        obs.append("Páginas de detalhe digitalizadas (imagem): UCs não extraídas.")
    elif not bate:
        obs.append(f"Soma das UCs ({_fmt_valor(soma)}) não bate com o total "
                   f"({_fmt_valor(total)}).")
    if contas is not None and not escaneado and len(unidades) != contas:
        obs.append(f"Contagem extraída ({len(unidades)}) difere do informado ({contas}).")

    bordero = {
        "id_bordero": id_bordero,
        "numero_fatura_agrupada": numero,
        "arquivo_pdf": nome,
        "distribuidora": distribuidora,
        "cod_agrupamento": cod_agr,
        "competencia": competencia,
        "data_vencimento": vencimento,
        "quantidade_contas": contas if contas is not None else "",
        "quantidade_contas_extraidas": len(unidades),
        "valor_total": total,
        "valor_total_bruto": bruto,
        "valor_total_retencoes": retencoes,
        "soma_valores_extraidos": soma,
        "bate_total": "SIM" if bate else ("N/A" if escaneado else "NÃO"),
        "escaneado": "SIM" if escaneado else "NÃO",
        "observacao": " ".join(obs),
    }

    # Carimba id/competência/distribuidora nas linhas de detalhe.
    for u in unidades:
        u.setdefault("nota_fiscal", "")
        for k in ("valor_bruto", "cofins", "pis", "irrf", "csll"):
            u.setdefault(k, None)
        u["id_bordero"] = id_bordero
        u["distribuidora"] = distribuidora
        u["competencia"] = competencia
    for r in resumo:
        r["id_bordero"] = id_bordero
        r["distribuidora"] = distribuidora
        r["competencia"] = competencia

    return ResultadoBordero(bordero=bordero, unidades=unidades, resumo=resumo)


def listar_pdfs(pasta: str) -> list[str]:
    """Lista, ordenados, todos os PDFs de borderô abaixo de `pasta` (recursivo)."""
    return sorted(glob.glob(os.path.join(pasta, "**", "*.pdf"), recursive=True))


def processar_pasta(pasta: str, progresso=None):
    """
    Processa todos os borderôs de `pasta`. `progresso(i, n, nome)` é chamado a
    cada arquivo, se fornecido. Devolve (borderos, unidades, resumo) — listas de
    dicts prontas para virar DataFrame/planilha.
    """
    pdfs = listar_pdfs(pasta)
    borderos, unidades, resumo = [], [], []
    for i, p in enumerate(pdfs, 1):
        if progresso:
            progresso(i, len(pdfs), os.path.basename(p))
        res = processar_pdf(p)
        borderos.append(res.bordero)
        unidades.extend(res.unidades)
        resumo.extend(res.resumo)
    return borderos, unidades, resumo


# --------------------------------------------------------------------------- #
# Escrita da planilha (estilizada, mesmo visual do restante do app)
# --------------------------------------------------------------------------- #
def _dfs(borderos, unidades, resumo):
    import pandas as pd

    dfb = pd.DataFrame(borderos, columns=BORDERO_COLS)
    dfu = pd.DataFrame(unidades, columns=UNIDADE_COLS)
    dfr = pd.DataFrame(resumo, columns=RESUMO_COLS)

    # Valores monetários como número (deixa o Excel formatar/somar).
    for col in ("valor_total", "valor_total_bruto", "valor_total_retencoes",
                "soma_valores_extraidos"):
        dfb[col] = pd.to_numeric(dfb[col], errors="coerce")
    for col in ("valor_bruto", "cofins", "pis", "irrf", "csll", "valor"):
        if col in dfu:
            dfu[col] = pd.to_numeric(dfu[col], errors="coerce")
    if "valor" in dfr:
        dfr["valor"] = pd.to_numeric(dfr["valor"], errors="coerce")
    return {ABA_BORDEROS: dfb, ABA_UNIDADES: dfu, ABA_RESUMO: dfr}


def _estilizar(caminho, abas):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(border_style="thin", color="BFBFBF")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = load_workbook(caminho)
    for sn in abas:
        if sn not in wb.sheetnames:
            continue
        hc, rc = SHEET_COLORS.get(sn, ("404040", "D9D9D9"))
        ws = wb[sn]
        hf = PatternFill("solid", fgColor=hc)
        rf = PatternFill("solid", fgColor=rc)
        for cell in ws[1]:
            cell.fill = hf
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = brd
        ws.row_dimensions[1].height = 30
        alt = False
        for row in ws.iter_rows(min_row=2):
            f = PatternFill("solid", fgColor="FFFFFF") if not alt else rf
            for c in row:
                c.fill = f
                c.font = Font(size=10)
                c.border = brd
                c.alignment = Alignment(vertical="center")
                if isinstance(c.value, float):
                    c.number_format = "#,##0.00"
            alt = not alt
        for col in ws.columns:
            mx = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 55)
        ws.freeze_panes = "A2"
    wb.save(caminho)


def escrever_dfs(dfs: dict, caminho: str):
    """Grava um conjunto arbitrário de DataFrames (já no layout das abas de
    borderô) em `caminho`, preservando a mesma formatação visual."""
    import pandas as pd

    abas = [a for a in SHEET_ORDER if a in dfs] + [a for a in dfs if a not in SHEET_ORDER]
    with pd.ExcelWriter(caminho, engine="openpyxl") as w:
        for aba in abas:
            df = dfs[aba]
            if df is None or df.empty:
                pd.DataFrame({"(sem dados)": []}).to_excel(w, sheet_name=aba, index=False)
            else:
                df.to_excel(w, sheet_name=aba, index=False)
    _estilizar(caminho, abas)
    return caminho


def escrever_planilha(borderos, unidades, resumo, caminho: str):
    """Grava as três abas (borderos/unidades/resumo_valores) em `caminho`."""
    dfs = _dfs(borderos, unidades, resumo)
    return escrever_dfs(dfs, caminho)


def gerar_planilha(pasta: str, caminho: str, progresso=None):
    """Conveniência: processa `pasta` e grava a planilha em `caminho`."""
    borderos, unidades, resumo = processar_pasta(pasta, progresso)
    escrever_planilha(borderos, unidades, resumo, caminho)
    return borderos, unidades, resumo


# --------------------------------------------------------------------------- #
# Leitura de uma planilha de borderôs já existente (sem metadados embutidos —
# o esquema de colunas é fixo, então não há renomeação/mapeamento a resolver).
# --------------------------------------------------------------------------- #
def ler_planilha(caminho: str) -> dict:
    """Lê todas as abas visíveis de `caminho`. Devolve {nome_aba: DataFrame}."""
    import pandas as pd

    xls = pd.ExcelFile(caminho, engine="openpyxl")
    dfs = {}
    for nome in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=nome, dtype=object)
        if list(df.columns) == ["(sem dados)"]:
            df = pd.DataFrame()
        dfs[nome] = df
    return dfs


# --------------------------------------------------------------------------- #
# "Adicionar a uma planilha" — complementar uma planilha de borderôs existente
# com novos PDFs processados (mesmo espírito da aba de faturas, mas com
# esquema de colunas FIXO, já que borderôs não suportam renomeação/exclusão de
# colunas pelo usuário).
# --------------------------------------------------------------------------- #
def divergencias(base_dfs: dict) -> list[str]:
    """
    Compara as abas/colunas da planilha BASE com o esquema fixo de borderôs.
    Devolve uma lista de mensagens (uma por divergência); vazia se compatível.
    """
    msgs: list[str] = []
    for aba, cols in ((ABA_BORDEROS, BORDERO_COLS), (ABA_UNIDADES, UNIDADE_COLS),
                     (ABA_RESUMO, RESUMO_COLS)):
        base_df = base_dfs.get(aba)
        if base_df is None:
            msgs.append(f"Aba '{aba}': inédita — não existe na planilha base.")
            continue
        faltantes = [c for c in cols if c not in base_df.columns]
        if faltantes:
            msgs.append(f"Aba '{aba}': coluna(s) inédita(s), ausente(s) na planilha "
                        f"base — {', '.join(faltantes)}.")
    return msgs


def concatenar(base_dfs: dict, novos_borderos: list, novos_unidades: list,
              novos_resumo: list):
    """
    Concatena o resultado de `processar_pasta` com uma planilha base já
    existente. Dedup: 'borderos' pela chave 'id_bordero' (evita duplicar um
    borderô já processado); 'unidades'/'resumo_valores' pela linha inteira
    (o mesmo bordê pode legitimamente repetir uma UC/rubrica).
    Devolve (dfs_resultado, resumo:list[str]).
    """
    import pandas as pd

    novos_dfs = _dfs(novos_borderos, novos_unidades, novos_resumo)
    resultado: dict = {}
    resumo: list[str] = []
    especs = ((ABA_BORDEROS, BORDERO_COLS, ["id_bordero"]),
              (ABA_UNIDADES, UNIDADE_COLS, None),
              (ABA_RESUMO, RESUMO_COLS, None))
    for aba, cols, chave in especs:
        base_df = base_dfs.get(aba)
        base_df = base_df.reindex(columns=cols) if base_df is not None else pd.DataFrame(columns=cols)
        novo_df = novos_dfs[aba].reindex(columns=cols)
        combinado = pd.concat([base_df, novo_df], ignore_index=True)
        antes = len(combinado)
        subset = chave if chave else None
        combinado = combinado.drop_duplicates(subset=subset, keep="first").reset_index(drop=True)
        n_dup = antes - len(combinado)
        resultado[aba] = combinado
        linha = (f"Aba '{aba}': {len(base_df)} (base) + {len(novo_df)} (novos) "
                f"= {len(combinado)} linha(s).")
        if n_dup:
            linha += f" ({n_dup} duplicata(s) removida(s))."
        resumo.append(linha)
    return resultado, resumo
