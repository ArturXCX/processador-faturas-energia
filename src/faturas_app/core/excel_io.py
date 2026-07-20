"""
Leitura e escrita das planilhas Excel.

- Escrita: aplica a formatação idêntica à dos notebooks (cabeçalho colorido,
  linhas zebradas, bordas, largura automática, painel congelado) e grava uma
  aba OCULTA de metadados (`_faturas_meta`) com o mapa nome_exibido -> canônico.
- Leitura: devolve os DataFrames das abas visíveis (como texto/objeto) e, se
  existir, os metadados embutidos.
"""
from __future__ import annotations

import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import schema


def _estilizar(caminho, abas_presentes):
    thin = Side(border_style="thin", color="BFBFBF")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = load_workbook(caminho)
    for sn in abas_presentes:
        if sn not in wb.sheetnames:
            continue
        hc, rc = schema.SHEET_COLORS.get(sn, ("404040", "D9D9D9"))
        ws = wb[sn]
        hf = PatternFill("solid", fgColor=hc)
        rf = PatternFill("solid", fgColor=rc)
        for cell in ws[1]:
            cell.fill = hf
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = brd
        ws.row_dimensions[1].height = 30
        alt = False
        for row in ws.iter_rows(min_row=2):
            f = PatternFill("solid", fgColor="FFFFFF") if not alt else rf
            for c in row:
                c.fill = f
                c.font = Font(size=10)
                c.border = brd
                c.alignment = Alignment(vertical="center")
            alt = not alt
        for col in ws.columns:
            mx = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 50)
        ws.freeze_panes = "A2"

    # Aba oculta de metadados (sem estilo).
    if schema.META_SHEET in wb.sheetnames:
        wb[schema.META_SHEET].sheet_state = "hidden"
    wb.save(caminho)


def escrever_workbook(display_dfs: dict[str, pd.DataFrame], meta: dict, caminho: str):
    """
    Grava os DataFrames de exibição em `caminho`, com formatação e metadados.
    `display_dfs` já deve estar com nomes exibidos e colunas/abas selecionadas.
    """
    abas_ordenadas = [a for a in schema.SHEET_ORDER if a in display_dfs]
    # Abas fora do schema padrão (ex.: vindas de um upload) entram ao final.
    abas_ordenadas += [a for a in display_dfs if a not in abas_ordenadas]

    with pd.ExcelWriter(caminho, engine="openpyxl") as w:
        for aba in abas_ordenadas:
            df = display_dfs[aba]
            if df is None or df.empty:
                pd.DataFrame({"(sem dados)": []}).to_excel(w, sheet_name=aba, index=False)
            else:
                df.to_excel(w, sheet_name=aba, index=False)
        # Metadados em aba própria (1 célula com JSON).
        meta_json = json.dumps(meta, ensure_ascii=False)
        pd.DataFrame({"meta": [meta_json]}).to_excel(
            w, sheet_name=schema.META_SHEET, index=False)

    _estilizar(caminho, abas_ordenadas)


def ler_workbook(caminho: str):
    """
    Lê todas as abas visíveis de `caminho`. Devolve (display_dfs, meta) onde
    `meta` é o dict de metadados embutidos, ou None se ausente.
    """
    xls = pd.ExcelFile(caminho, engine="openpyxl")
    display_dfs: dict[str, pd.DataFrame] = {}
    meta = None
    for nome in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=nome, dtype=object)
        if nome == schema.META_SHEET:
            try:
                meta = json.loads(str(df.iloc[0, 0]))
            except Exception:
                meta = None
            continue
        # Ignora abas "(sem dados)" placeholder.
        if list(df.columns) == ["(sem dados)"]:
            df = pd.DataFrame()
        display_dfs[nome] = df
    return display_dfs, meta
