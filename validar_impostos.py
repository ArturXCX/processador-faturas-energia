#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Reconciliação de impostos das faturas de energia.

Confere, por fatura, se o total de tributos da aba `impostos` bate com a soma
dos tributos distribuídos na aba `itens_fatura`, e regenera o relatório de
pendências em ``planilhas/pendencia/impostos_faturas.xlsx``.

Dois cuidados que a verificação anterior não tinha e que geravam falsos
positivos:

1. TOLERÂNCIA DE ARREDONDAMENTO. A aba `impostos` traz o total OFICIAL,
   consolidado, de cada tributo (PIS, COFINS, ICMS). A soma item a item usa
   valores já arredondados ao centavo, então acumula alguns centavos. Isso é
   ruído, não erro — só é divergência real acima de ``TOLERANCIA`` (R$ 0,50),
   que separa com folga o arredondamento (tudo <= R$ 0,30 no acervo) dos
   poucos casos materiais.

2. ITEM CONSOLIDADO "TRIBUTOS (PIS/COFINS)". Em faturas CHESP no layout antigo
   ("Modelo 6"), o PIS+COFINS NÃO é distribuído por item — aparece como um
   único item ``TRIBUTOS (PIS/COFINS)`` cujo ``valor_r$`` é o total. Somar só a
   coluna ``pis_cofins`` dava zero e acusava a fatura inteira como divergente.
   Aqui esse item entra na soma dos itens.

Uso:
    python validar_impostos.py [caminho_v9.xlsx]
"""
from __future__ import annotations

import sys
from collections import defaultdict

import openpyxl

V9_PADRAO = "resultados/faturas_energia_v9.xlsx"
SAIDA = "planilhas/pendencia/impostos_faturas.xlsx"
TOLERANCIA = 0.50  # R$ — acima disso deixa de ser arredondamento

# Itens cujo valor_r$ representa PIS+COFINS consolidado (não vão na coluna
# pis_cofins). Mantido como tupla para comparação exata do nome normalizado.
ITENS_TRIBUTO_CONSOLIDADO = ("TRIBUTOS (PIS/COFINS)",)


def _rows(wb, sheet):
    ws = wb[sheet]
    hdr = next(ws.iter_rows(values_only=True))
    for r in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(hdr, r))


def reconciliar(v9_path: str):
    wb = openpyxl.load_workbook(v9_path, read_only=True)

    # total da aba impostos por fatura
    aba_impostos = defaultdict(float)
    for d in _rows(wb, "impostos"):
        aba_impostos[d["id_fatura"]] += (d.get("Valor (R$)") or 0)

    # soma dos tributos nos itens: pis_cofins + icms + itens consolidados
    soma_itens = defaultdict(float)
    total_faturas = defaultdict(float)
    valor_total = {}
    ids = set(aba_impostos)
    for d in _rows(wb, "itens_fatura"):
        fid = d["id_fatura"]
        ids.add(fid)
        item = str(d.get("item") or "")
        total_faturas[fid] += (d.get("valor_r$") or 0)
        if item in ITENS_TRIBUTO_CONSOLIDADO:
            soma_itens[fid] += (d.get("valor_r$") or 0)
        else:
            soma_itens[fid] += (d.get("pis_cofins") or 0) + (d.get("icms") or 0)
    for d in _rows(wb, "fatura"):
        valor_total[d["id_fatura"]] = d.get("valor_total_r$")
        ids.add(d["id_fatura"])
    wb.close()

    div_impostos = []   # (id, aba, soma_itens, diferenca)
    div_total = []      # (id, valor_total, soma_itens_valor, diferenca)
    for fid in sorted(ids):
        a = round(aba_impostos.get(fid, 0), 2)
        s = round(soma_itens.get(fid, 0), 2)
        if abs(a - s) > TOLERANCIA:
            div_impostos.append((fid, a, s, round(a - s, 2)))
        vt = valor_total.get(fid)
        st = round(total_faturas.get(fid, 0), 2)
        if vt is not None and abs((vt or 0) - st) > TOLERANCIA:
            div_total.append((fid, round(vt, 2), st, round((vt or 0) - st, 2)))

    return len(ids), div_total, div_impostos


def escrever_relatorio(total, div_total, div_impostos, saida: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Verificacao", "Total", "Inconsistencias", "Tolerancia (R$)"])
    ws.append(["Total fatura = soma itens", total, len(div_total), TOLERANCIA])
    ws.append(["Impostos = soma impostos itens", total, len(div_impostos), TOLERANCIA])

    ws = wb.create_sheet("Impostos_divergencias")
    ws.append(["id_fatura", "aba_impostos", "soma_itens", "diferenca"])
    for row in sorted(div_impostos, key=lambda x: -abs(x[3])):
        ws.append(list(row))

    if div_total:
        ws = wb.create_sheet("Total_divergencias")
        ws.append(["id_fatura", "valor_total", "soma_itens", "diferenca"])
        for row in sorted(div_total, key=lambda x: -abs(x[3])):
            ws.append(list(row))

    wb.save(saida)


def main():
    v9 = sys.argv[1] if len(sys.argv) > 1 else V9_PADRAO
    total, div_total, div_impostos = reconciliar(v9)
    escrever_relatorio(total, div_total, div_impostos, SAIDA)
    print(f"Faturas verificadas: {total}")
    print(f"Tolerancia: R$ {TOLERANCIA:.2f}")
    print(f"Divergencias 'fatura = soma itens': {len(div_total)}")
    print(f"Divergencias 'impostos = soma itens': {len(div_impostos)}")
    for row in sorted(div_impostos, key=lambda x: -abs(x[3])):
        print(f"   {row[0]:28} aba={row[1]:>10.2f} itens={row[2]:>10.2f} dif={row[3]:>8.2f}")
    print(f"Relatorio: {SAIDA}")


if __name__ == "__main__":
    main()
